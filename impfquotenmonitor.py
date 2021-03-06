import openpyxl
from pathlib import Path
import requests
from bs4 import BeautifulSoup
import urllib.parse
import os,sys,re,random
from datetime import datetime
from SPARQLWrapper import SPARQLWrapper, JSON
import pystache

def get_spreadsheet():
    filename = 'Impfquotenmonitoring.xlsx'
    url = "https://www.rki.de/DE/Content/InfAZ/N/Neuartiges_Coronavirus/Daten/"
    url += filename
    url += "?__blob=publicationFile"
    r = requests.get(url, allow_redirects=True)
    open(filename, 'wb').write(r.content)
    return filename

def get_sum():
    xlsx_file = Path(os.getcwd(), get_spreadsheet())
    wb_obj = openpyxl.load_workbook(xlsx_file)
    sheet = wb_obj[wb_obj.sheetnames[1]]
    first_jab = 0
    second_jab = 0
    for i in range(4,20):
        if type(sheet.cell(row = i, column = 4).value) == int:
            first_jab += sheet.cell(row = i, column = 4).value
        if type(sheet.cell(row = i, column = 9).value) == int:
            second_jab += sheet.cell(row = i, column = 9).value
    return first_jab, second_jab

def get_status():
    url = "https://www.rki.de/DE/Content/InfAZ/N/Neuartiges_Coronavirus/Daten"
    url += "/Impfquoten-Tab.html"
    r = requests.get(url, allow_redirects=True)
    soup = BeautifulSoup(r.content, 'html.parser')
    datenstand = soup.find_all(string=re.compile('Stand'))[0]
    return datenstand

def run_query(sum):
    diff = 1000
    result = []
    #query +=  "         wdt:P17 wd:Q183;\n"
    while True:
        endpoint_url = "https://query.wikidata.org/sparql"
        query = "SELECT DISTINCT ?city ?cityLabel ?population ?sitelink WHERE {\n"
        query +=  "?city wdt:P31/wdt:P279* wd:Q486972;\n"
        query +=  "         wdt:P1082 ?population.\n"
        query +=  "?sitelink schema:about ?city;\n"
        query +=  "          schema:isPartOf <https://de.wikipedia.org/>;\n"
        query +=  "FILTER (abs(?population - "+str(sum)+") < "+str(diff)+")\n"
        query +=  "SERVICE wikibase:label "
        query += '{ bd:serviceParam wikibase:language "[AUTO_LANGUAGE],de" }'
        query +="\n}"
        user_agent = "Impfquotenmonitorvergleich"
        sparql = SPARQLWrapper(endpoint_url, agent=user_agent)
        sparql.setQuery(query)
        sparql.setReturnFormat(JSON)
        results = sparql.query().convert()
        result = results["results"]["bindings"]
        if result:
            break
        diff += 500
    result = random.choice(results["results"]["bindings"])
    city = result["cityLabel"]["value"]
    sitelink = result["sitelink"]["value"]
    querylink = "https://query.wikidata.org/#" + urllib.parse.quote(query)
    return city, sitelink, querylink

if __name__ == "__main__":
    first_jab, second_jab = get_sum()
    city, sitelink, querylink = run_query(first_jab)
    datenstand = get_status()
    renderer = pystache.Renderer()
    with open("index.html", "w") as index_file:
        index_file.write(renderer.render_path(
            'index.mustache',
            {
                "first_jab": f'{first_jab:,}'.replace(',','&thinsp;'),
                "second_jab": f'{second_jab:,}'.replace(',','&thinsp;'),
                "city": city,
                "datenstand": datenstand,
                "sitelink": sitelink,
                "querylink": querylink
            }
        ))
